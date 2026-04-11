#!/usr/bin/env python3
"""
md2excel - Markdown を Excel (.xlsx) に変換する汎用CLIツール。

使い方:
  # 1ファイル → 1 Excel
  python md2excel.py docs/design/architecture.md

  # 複数ファイル → 1 Excel（シート分割）
  python md2excel.py docs/design/*.md --merge -o design-docs.xlsx

  # 出力先ディレクトリ指定
  python md2excel.py docs/design/db-design.md -o output/

対応要素:
  - 見出し (H1-H3)
  - テーブル (パイプ区切り)
  - リスト (箇条書き・番号付き)
  - コードブロック (Mermaid は画像埋め込み対応)
  - 段落テキスト
"""

import argparse
import atexit
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("エラー: openpyxl が必要です。インストールしてください:")
    print("  pip install openpyxl")
    sys.exit(1)

# Mermaid 画像の一時ディレクトリ（プロセス終了時に自動削除）
_mermaid_tmpdir: Path | None = None


def _get_mermaid_tmpdir() -> Path:
    global _mermaid_tmpdir
    if _mermaid_tmpdir is None:
        _mermaid_tmpdir = Path(tempfile.mkdtemp(prefix="md2excel-mermaid-"))
        atexit.register(lambda: shutil.rmtree(_mermaid_tmpdir, ignore_errors=True))
    return _mermaid_tmpdir


_mermaid_counter = 0


def render_mermaid_to_png(mermaid_code: str) -> Path | None:
    """Mermaid コードを PNG 画像にレンダリングする。失敗時は None を返す。"""
    global _mermaid_counter
    _mermaid_counter += 1
    tmpdir = _get_mermaid_tmpdir()
    input_file = tmpdir / f"diagram_{_mermaid_counter}.mmd"
    output_file = tmpdir / f"diagram_{_mermaid_counter}.png"

    input_file.write_text(mermaid_code, encoding="utf-8")

    try:
        subprocess.run(
            [
                "npx", "--yes", "@mermaid-js/mermaid-cli",
                "-i", str(input_file),
                "-o", str(output_file),
                "-b", "white",
                "--scale", "2",
            ],
            capture_output=True,
            timeout=30,
            check=True,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired, FileNotFoundError) as e:
        print(f"  警告: Mermaid レンダリング失敗（テキスト表示にフォールバック）: {e}", file=sys.stderr)
        return None

    if output_file.exists():
        return output_file
    return None


# ---------------------------------------------------------------------------
# スタイル定義
# ---------------------------------------------------------------------------

FONT_DEFAULT = Font(name="Meiryo", size=10)
FONT_H1 = Font(name="Meiryo", size=14, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Meiryo", size=12, bold=True, color="333333")
FONT_H3 = Font(name="Meiryo", size=11, bold=True, color="333333")
FONT_CODE = Font(name="Consolas", size=9, color="333333")
FONT_TABLE_HEADER = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
FONT_TABLE_CELL = Font(name="Meiryo", size=10)

FILL_H1 = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_H2 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_H3 = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
FILL_CODE = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_TABLE_HEADER = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
FILL_TABLE_ALT = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# 印刷幅の基準列数（A〜F の6列をマージ領域とする）
MERGE_COLS = 6


# ---------------------------------------------------------------------------
# Markdown パーサー
# ---------------------------------------------------------------------------


def parse_markdown(text: str) -> list[dict]:
    """Markdown テキストを構造化ブロックのリストに変換する。

    Returns:
        list of dicts。各 dict は type キーを持つ:
        - heading: level, text
        - table: headers (list[str]), rows (list[list[str]])
        - code: lang, text
        - list: items (list[str])
        - paragraph: text
    """
    lines = text.split("\n")
    blocks: list[dict] = []
    i = 0

    while i < len(lines):
        line = lines[i]

        # --- 空行スキップ ---
        if not line.strip():
            i += 1
            continue

        # --- コードブロック ---
        m = re.match(r"^```(\w*)", line)
        if m:
            lang = m.group(1) or ""
            code_lines = []
            i += 1
            while i < len(lines) and not lines[i].startswith("```"):
                code_lines.append(lines[i])
                i += 1
            i += 1  # 閉じ ``` をスキップ
            blocks.append({"type": "code", "lang": lang, "text": "\n".join(code_lines)})
            continue

        # --- 見出し ---
        m = re.match(r"^(#{1,3})\s+(.+)$", line)
        if m:
            level = len(m.group(1))
            blocks.append({"type": "heading", "level": level, "text": m.group(2).strip()})
            i += 1
            continue

        # --- テーブル ---
        if "|" in line and line.strip().startswith("|"):
            table_lines = []
            while i < len(lines) and "|" in lines[i] and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1
            blocks.append(_parse_table(table_lines))
            continue

        # --- リスト ---
        if re.match(r"^\s*[-*+]\s+", line) or re.match(r"^\s*\d+\.\s+", line):
            items = []
            while i < len(lines) and (
                re.match(r"^\s*[-*+]\s+", lines[i])
                or re.match(r"^\s*\d+\.\s+", lines[i])
            ):
                item_text = re.sub(r"^\s*[-*+]\s+", "", lines[i])
                item_text = re.sub(r"^\s*\d+\.\s+", "", item_text)
                items.append(item_text.strip())
                i += 1
            blocks.append({"type": "list", "items": items})
            continue

        # --- 段落 ---
        para_lines = []
        while i < len(lines) and lines[i].strip() and not _is_block_start(lines[i]):
            para_lines.append(lines[i].strip())
            i += 1
        if para_lines:
            blocks.append({"type": "paragraph", "text": " ".join(para_lines)})

    return blocks


def _is_block_start(line: str) -> bool:
    """次のブロック開始かどうかを判定する。"""
    if re.match(r"^#{1,3}\s+", line):
        return True
    if re.match(r"^```", line):
        return True
    if line.strip().startswith("|") and "|" in line:
        return True
    if re.match(r"^\s*[-*+]\s+", line) or re.match(r"^\s*\d+\.\s+", line):
        return True
    return False


def _parse_table(lines: list[str]) -> dict:
    """パイプ区切りテーブルをパースする。"""

    def split_row(line: str) -> list[str]:
        cells = line.strip().strip("|").split("|")
        return [c.strip() for c in cells]

    if len(lines) < 2:
        return {"type": "table", "headers": [], "rows": [split_row(lines[0])]}

    headers = split_row(lines[0])

    # セパレータ行（---）をスキップ
    start = 1
    if re.match(r"^[\s|:-]+$", lines[1]):
        start = 2

    rows = [split_row(l) for l in lines[start:]]
    return {"type": "table", "headers": headers, "rows": rows}


# ---------------------------------------------------------------------------
# インラインMarkdown除去
# ---------------------------------------------------------------------------


def _strip_inline_md(text: str) -> str:
    """インラインMarkdown記法（太字、コード、リンク等）をプレーンテキストに変換する。"""
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)  # **bold**
    text = re.sub(r"\*(.+?)\*", r"\1", text)  # *italic*
    text = re.sub(r"`(.+?)`", r"\1", text)  # `code`
    text = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", text)  # [text](url)
    return text


# ---------------------------------------------------------------------------
# Excel 書き出し
# ---------------------------------------------------------------------------


def write_blocks_to_sheet(ws, blocks: list[dict], title: str = "", render_mermaid: bool = True):
    """パース済みブロックを Excel シートに書き出す。"""
    row = 1

    # 列幅の初期設定
    col_widths = [30] * MERGE_COLS
    for c in range(1, MERGE_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths[c - 1]

    for block in blocks:
        btype = block["type"]

        if btype == "heading":
            row = _write_heading(ws, row, block)

        elif btype == "table":
            row = _write_table(ws, row, block)

        elif btype == "code":
            row = _write_code(ws, row, block, render_mermaid=render_mermaid)

        elif btype == "list":
            row = _write_list(ws, row, block)

        elif btype == "paragraph":
            row = _write_paragraph(ws, row, block)

        # ブロック間に空行を入れる
        row += 1

    # 印刷設定
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _write_heading(ws, row: int, block: dict) -> int:
    level = block["level"]
    text = _strip_inline_md(block["text"])

    cell = ws.cell(row=row, column=1, value=text)

    if level == 1:
        cell.font = FONT_H1
        cell.fill = FILL_H1
    elif level == 2:
        cell.font = FONT_H2
        cell.fill = FILL_H2
    else:
        cell.font = FONT_H3
        cell.fill = FILL_H3

    cell.alignment = Alignment(vertical="center")

    # 行をマージ
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=MERGE_COLS
    )

    # マージ範囲のセルにも背景色を適用
    if level == 1:
        fill = FILL_H1
    elif level == 2:
        fill = FILL_H2
    else:
        fill = FILL_H3
    for c in range(2, MERGE_COLS + 1):
        ws.cell(row=row, column=c).fill = fill

    ws.row_dimensions[row].height = {1: 30, 2: 25, 3: 22}.get(level, 20)

    return row + 1


def _write_table(ws, row: int, block: dict) -> int:
    headers = block["headers"]
    rows_data = block["rows"]
    num_cols = max(len(headers), max((len(r) for r in rows_data), default=0)) if headers or rows_data else 1

    # 列幅を内容に合わせて調整
    all_rows = ([headers] if headers else []) + rows_data
    for ci in range(num_cols):
        max_len = 8  # 最小幅
        for r in all_rows:
            if ci < len(r):
                text = _strip_inline_md(r[ci])
                # 日本語は1文字=約2幅として概算
                text_width = sum(2 if ord(ch) > 127 else 1 for ch in text)
                max_len = max(max_len, text_width)
        col_letter = get_column_letter(ci + 1)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # ヘッダー行
    if headers:
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=_strip_inline_md(h))
            cell.font = FONT_TABLE_HEADER
            cell.fill = FILL_TABLE_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        row += 1

    # データ行
    for ri, data_row in enumerate(rows_data):
        for ci, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=ci, value=_strip_inline_md(val))
            cell.font = FONT_TABLE_CELL
            cell.alignment = ALIGN_WRAP
            cell.border = THIN_BORDER
            if ri % 2 == 1:
                cell.fill = FILL_TABLE_ALT
        row += 1

    return row


def _write_code(ws, row: int, block: dict, render_mermaid: bool = True) -> int:
    lang = block["lang"]
    text = block["text"]

    # Mermaid ブロック → 画像埋め込みを試行
    if lang == "mermaid" and render_mermaid:
        png_path = render_mermaid_to_png(text)
        if png_path:
            return _write_mermaid_image(ws, row, png_path)

    # 通常のコードブロック（または Mermaid レンダリング失敗時のフォールバック）
    if lang:
        label = f"[{lang}]"
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name="Consolas", size=9, bold=True, color="666666")
        cell.fill = FILL_CODE
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=MERGE_COLS
        )
        for c in range(2, MERGE_COLS + 1):
            ws.cell(row=row, column=c).fill = FILL_CODE
        row += 1

    for code_line in text.split("\n"):
        cell = ws.cell(row=row, column=1, value=code_line)
        cell.font = FONT_CODE
        cell.fill = FILL_CODE
        cell.alignment = Alignment(wrap_text=False, vertical="top")
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=MERGE_COLS
        )
        for c in range(2, MERGE_COLS + 1):
            ws.cell(row=row, column=c).fill = FILL_CODE
        row += 1

    return row


def _write_mermaid_image(ws, row: int, png_path: Path) -> int:
    """Mermaid レンダリング結果の PNG 画像をシートに埋め込む。"""
    img = XlImage(str(png_path))

    # 画像サイズをシート幅に収める（6列 x 約70px ≒ 420px を基準にスケーリング）
    max_width_px = MERGE_COLS * 70
    if img.width > max_width_px:
        scale = max_width_px / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)

    # 画像が占める行数を概算（1行 ≒ 15px）
    row_height_px = 15
    rows_needed = max(1, img.height // row_height_px + 1)

    anchor = f"A{row}"
    ws.add_image(img, anchor)

    return row + rows_needed


def _write_list(ws, row: int, block: dict) -> int:
    for idx, item in enumerate(block["items"], 1):
        text = _strip_inline_md(item)
        cell = ws.cell(row=row, column=1, value=f"  {idx}. {text}")
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_WRAP
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=MERGE_COLS
        )
        row += 1
    return row


def _write_paragraph(ws, row: int, block: dict) -> int:
    text = _strip_inline_md(block["text"])
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_DEFAULT
    cell.alignment = ALIGN_WRAP
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=MERGE_COLS
    )
    return row + 1


# ---------------------------------------------------------------------------
# ファイル変換
# ---------------------------------------------------------------------------


def convert_file(md_path: Path) -> list[dict]:
    """Markdown ファイルを読み込んでパース済みブロックを返す。"""
    text = md_path.read_text(encoding="utf-8")
    return parse_markdown(text)


def make_sheet_name(path: Path) -> str:
    """ファイルパスからExcelシート名（31文字以内）を生成する。"""
    name = path.stem
    # シート名に使えない文字を除去
    name = re.sub(r'[\\/*?\[\]:]', "", name)
    return name[:31]


def convert_single(md_path: Path, output_path: Path, render_mermaid: bool = True) -> Path:
    """1 MD → 1 XLSX 変換。"""
    blocks = convert_file(md_path)
    wb = Workbook()
    ws = wb.active
    ws.title = make_sheet_name(md_path)
    write_blocks_to_sheet(ws, blocks, title=md_path.stem, render_mermaid=render_mermaid)
    wb.save(str(output_path))
    return output_path


def convert_merged(md_paths: list[Path], output_path: Path, render_mermaid: bool = True) -> Path:
    """複数 MD → 1 XLSX（シート分割）変換。"""
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    for md_path in md_paths:
        blocks = convert_file(md_path)
        sheet_name = make_sheet_name(md_path)

        # 重複シート名の回避
        existing = {ws.title for ws in wb.worksheets}
        if sheet_name in existing:
            for suffix in range(2, 100):
                candidate = f"{sheet_name[:28]}_{suffix}"
                if candidate not in existing:
                    sheet_name = candidate
                    break

        ws = wb.create_sheet(title=sheet_name)
        write_blocks_to_sheet(ws, blocks, title=md_path.stem, render_mermaid=render_mermaid)

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Markdown を Excel (.xlsx) に変換する",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  # 1ファイルずつ変換
  python md2excel.py docs/design/architecture.md

  # 複数ファイルを1ブックにまとめる
  python md2excel.py docs/design/*.md --merge -o design-docs.xlsx

  # 出力先ディレクトリ指定（ファイル名は自動）
  python md2excel.py docs/design/db-design.md -o output/
        """,
    )
    parser.add_argument(
        "files",
        nargs="+",
        type=Path,
        help="変換対象の Markdown ファイル（複数指定可）",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="出力先パス（ファイル名 or ディレクトリ）",
    )
    parser.add_argument(
        "--merge",
        action="store_true",
        help="複数ファイルを1つの Excel ブック（シート分割）にまとめる",
    )
    parser.add_argument(
        "--no-mermaid",
        action="store_true",
        help="Mermaid 図の画像レンダリングを無効化し、コードテキストで表示する",
    )

    args = parser.parse_args()
    render_mermaid = not args.no_mermaid

    # 入力ファイルの存在チェック
    md_files: list[Path] = []
    for f in args.files:
        if not f.exists():
            print(f"エラー: ファイルが見つかりません: {f}", file=sys.stderr)
            sys.exit(1)
        if not f.is_file():
            print(f"エラー: ファイルではありません: {f}", file=sys.stderr)
            sys.exit(1)
        md_files.append(f.resolve())

    if not md_files:
        print("エラー: 変換対象のファイルがありません", file=sys.stderr)
        sys.exit(1)

    # マージモード
    if args.merge:
        if args.output:
            out = args.output
            if out.is_dir():
                out = out / "merged.xlsx"
            elif not out.suffix:
                out = out.with_suffix(".xlsx")
        else:
            out = Path("merged.xlsx")

        out.parent.mkdir(parents=True, exist_ok=True)
        result = convert_merged(md_files, out, render_mermaid=render_mermaid)
        print(f"生成: {result}  ({len(md_files)} シート)")

    # 個別変換モード
    else:
        for md_file in md_files:
            if args.output:
                out = args.output
                # ディレクトリ判定: 既存ディレクトリ or 拡張子なし or 複数ファイル指定
                is_dir = out.is_dir() or out.suffix == "" or len(md_files) > 1
                if is_dir:
                    out.mkdir(parents=True, exist_ok=True)
                    out = args.output / md_file.with_suffix(".xlsx").name
            else:
                out = md_file.with_suffix(".xlsx")

            out.parent.mkdir(parents=True, exist_ok=True)
            result = convert_single(md_file, out, render_mermaid=render_mermaid)
            print(f"生成: {result}")


if __name__ == "__main__":
    main()
